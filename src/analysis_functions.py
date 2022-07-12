import numpy as np
import pyabf
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def c1_vm(sweep_y_data, c1):
    """Returns the Vm at the cursor 1 position.

    Keyword arguments:
    sweep_y_data -- sweep Vm values
    c1 -- cursor 1 position in ms
    """
    return sweep_y_data[c1]


def sweep_command(abf_data):
    """Accesses the command epoch table in abf_data (current steps in pA),
    returns a list of current injections of the first 5 sweeps."""

    abf = abf_data
    # Each current step decreases by 1/4 of the initial step
    multiplier = [1, 0.75, 0.5, 0.25, 0]
    # Epochs.level index 2 gives the value of the initial command current
    current_injection = abf.sweepEpochs.levels[2]

    return [current_injection * x / 1000 for x in multiplier]  # /1000 to get nA


def c1c2_min(sweep_y_data, c1, c2):
    """Returns the lowest point between cursor 1 and 2.

    Keyword arguments:
    sweep_y_data -- sweep Vm values
    c1 -- cursor 1 position in ms
    c2 -- cursor 2 position in ms
    """
    return min(sweep_y_data[c1:c2])


def c3c4_min(sweep_y_data, c3, c4):
    """Returns the lowest point between cursor 3 and 4.

    Keyword arguments:
    sweep_y_data -- sweep Vm values
    c3 -- cursor 3 position in ms
    c4 -- cursor 4 position in ms
    """
    return min(sweep_y_data[c3:c4])


def peaks(sweep_y_data, threshold):
    """Returns peak indices for each sweep.

    Keyword arguments:
    sweep_x_data -- sweep time values
    sweep_y_data -- sweep Vm values
    threshold -- minimum detection threshold (in mV)
    """

    # Array of all sweep_y_data[i+1] - sweep_y_data[i]
    dvdt = np.diff(sweep_y_data)
    threshold_filter = np.where(sweep_y_data > threshold)[0]
    peak_indices = [i
                    for i in threshold_filter
                    if np.all(dvdt[i: i + 3] <= 0)
                    and np.all(dvdt[i - 10: i] > 0)
                    ]
    return peak_indices


def file_ids(day, start, end):
    """Returns the iv and cc file_id information.

    Keyword arguments:
    day -- day of recording
    start -- initial iv file
    end -- final iv file + 1
    """

    # Array of even and odd numbers between start and end (with a leading 0 below 10)
    file_ids_even = [
        str(x) if x > 9 else "0" + str(x) for x in range(start, end) if x % 2 == 0
    ]
    file_ids_odd = [
        str(x) if x > 9 else "0" + str(x) for x in range(start, end) if x % 2 == 1
    ]

    # If start is even then iv_file_ids = file_ids_even
    iv_file_ids = file_ids_even if start % 2 == 0 else file_ids_odd
    iv_files = [day + "_00{}".format(value) for value in iv_file_ids]

    cc_file_ids = file_ids_odd if start % 2 == 0 else file_ids_even
    cc_files = [day + "_00{}".format(value) for value in cc_file_ids]

    return [iv_files, cc_files]


def iv_plots(iv_files, current_cell, day, data_folder):
    """Creates the IV plots.

    Keyword arguments:
    iv_files -- iv_file codes
    day -- the day of recording
    data_folder -- path to Raw data folder
    """

    #  use a custom colourmap to create a different colour for every sweep
    #  add name of colourmap in: cmap = mpl.cm."HERE"
    #  (e.g. viridis,winter,rainbow,magma,BuGn,tab10,cividis)
    #  cmap = colourmap as a matrix of colours
    #  and then select a part of this matrix as your new colourmap
    #  change what sweeps are coloured in: colours = [cmap(x / "HERE")
    #  (e.g. len(iv_files) or abf.sweepCount)

    # cmap = mpl.cm.tab10(np.linspace(0, 1, 20))
    # cmap = mpl.colors.ListedColormap(cmap[0:, :-1])  # change first value from 0-19
    # colours = [cmap(x / len(iv_files)) for x in abf.sweepList]
    # colours.reverse()

    colours = [
        "tab:blue",
        "tab:orange",
        "grey",
        "goldenrod",
        "tab:cyan",
        "tab:green",
        "navy",
    ]

    for count, iv_file_name in enumerate(iv_files):
        path = data_folder + day + "/" + iv_file_name + ".abf"
        abf = pyabf.ABF(path)  # One IV file

        # fig = plt.figure(figsize=(8, 10))
        single_colour = colours[count]

        plt.title(f"Cell {current_cell} IV ({iv_file_name})")
        plt.ylabel(abf.sweepLabelY)
        plt.xlabel(abf.sweepLabelX)

        for i in range(abf.sweepCount):
            abf.setSweep(i)  # One sweep

            plt.plot(
                abf.sweepX,
                abf.sweepY,
                alpha=1,
                # label="sweep {}".format(i + 1),
                color=single_colour,
            )
            # if i >= 5:  # ignore first 5 traces
            # peak_i = peaks(abf.sweepX, abf.sweepY, threshold)
            # plt.plot(abf.sweepX[peak_i], abf.sweepY[peak_i], 'ro')
        plt.xlim([0, 2])
        # plt.legend()

        plt.savefig(
            data_folder + day + "/" + "IV " + iv_file_name, dpi=300, bbox_inches="tight"
        )
        plt.show()

    return


def cc_plots(cc_files, current_cell, day, data_folder):
    """Creates the CC plots.

    Keyword arguments:
    cc_files -- cc_file codes
    day -- the day of recording
    data_folder -- path to Raw data folder
    """

    for count, cc_file_name in enumerate(cc_files):
        path = data_folder + day + "/" + cc_file_name + ".abf"
        abf = pyabf.ABF(path)

        # fig = plt.figure(figsize=(15, 10))

        plt.title(f"Cell {current_cell} CC ({cc_file_name})")
        plt.ylabel(abf.sweepLabelY)
        plt.xlabel(abf.sweepLabelX)

        for i in range(abf.sweepCount):
            abf.setSweep(i, absoluteTime=True)  # <-- relates to sweepX
            plt.plot(
                abf.sweepX,
                abf.sweepY,
                alpha=1,
                # label="sweep {}".format(i + 1),
                color="black",
            )
        plt.ylim([-90, -0])
        # plt.legend()

        for i, tagTimeSec in enumerate(abf.tagTimesSec):
            pos_x = abf.tagTimesSec[i]
            comment = abf.tagComments[i]
            color = "C{}".format(i + 2)
            plt.axvline(pos_x, label=comment, color=color, ls='--')
        plt.legend(prop={'size': 25}, loc='upper right')

        plt.savefig(data_folder + day + "/" + "CC " + cc_file_name, dpi=300, bbox_inches="tight")
        plt.show()

    return


def iv_analysis(iv_files, current_cell, day, data_folder):
    """Runs the IV analysis and saves the results to individual Excel sheets.

    Keyword arguments:
    iv_files -- iv_file codes
    current_cell -- the current cell being analysed
    day -- the day of recording
    data_folder -- path to Raw data folder
    """
    wb = load_workbook(filename=data_folder + "file_ids.xlsx")

    cursor_1 = 298  # Cursor position in ms. Values obtained from Excel.
    cursor_2 = 4181
    cursor_3 = 8912
    cursor_4 = 10747
    threshold = 10  # detection threshold for peaks

    # ws = wb.active
    ws = wb.create_sheet(f"Cell {current_cell}", current_cell)  # Create a sheet for each cell
    for count, iv_file_name in enumerate(iv_files):
        path = data_folder + day + "/" + iv_file_name + ".abf"
        abf = pyabf.ABF(path)  # One IV file

        rmp = []
        current_injection = sweep_command(abf)
        ih = []
        ss = []
        peak_i = []
        sweep_first_ap = "No AP"
        for i in range(abf.sweepCount):
            abf.setSweep(i)  # One sweep
            if i <= 4:  # sweeps between 1-5
                rmp.append(c1_vm(abf.sweepY, cursor_1))
                ih.append(c1c2_min(abf.sweepY, cursor_1, cursor_2))
                ss.append(c3c4_min(abf.sweepY, cursor_3, cursor_4))
            elif peak_i == []:  # sweeps > 5. Finds the first peaks in the IV recording
                peak_i = peaks(abf.sweepY, threshold)

            if sweep_first_ap == "No AP" and peak_i != []:
                sweep_first_ap = i + 1
                print("IV {}.".format(iv_file_name), "Sweep with first AP:", sweep_first_ap)

        ws.cell(row=29, column=count * 5 + 2, value=sweep_first_ap)  # Sweep with first AP
        ws.cell(row=2, column=count * 5 + 2, value=int(iv_file_name[-2:]))  # IV file numbers

        results = [rmp, current_injection, ih, ss]
        for col, data in enumerate(results):
            for j, datum in enumerate(data):
                ws.cell(row=j + 5, column=col + 1 + count * 5, value=datum)

    # Titles
    ws.cell(row=1, column=2, value=day)
    ws.cell(row=2, column=1, value="Files")
    ws.cell(row=29, column=1, value="First AP")

    wb.save(
        filename=data_folder + "Results_copy.xlsx"
    )

    return
